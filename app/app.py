# dashboard_multi_platform_streamlit.py
# Gabungan 3 tools: Shopee & CPAS, META, TikTok
# Didesain agar masing-masing app bisa diakses tanpa mengubah logika aslinya.

import streamlit as st
import pandas as pd
import numpy as np
import io
import os
import re
from io import BytesIO
from datetime import datetime, date
from typing import Optional
from collections import OrderedDict
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from pandas.io.formats.style import Styler

# Set global page config once
st.set_page_config(page_title="Multi-Platform Excel Utilities", layout="wide")

# -----------------------------
# NAVBAR (Top horizontal) — pilih halaman platform
# -----------------------------
PAGES = ["Shopee", "Meta", "TikTok"]

# 1. Inisialisasi awal session state
if "page" not in st.session_state:
    st.session_state.page = PAGES[0]

# 2. Buat fungsi callback untuk tombol navbar
def set_page(selected_page):
    st.session_state.page = selected_page

def navbar():
    cols = st.columns(len(PAGES), gap="small")
    for i, p in enumerate(PAGES):
        with cols[i]:
            # 3. Gunakan on_click agar state berubah SEBELUM UI di-render ulang
            st.button(p, key=f"nav_{i}", on_click=set_page, args=(p,))
    st.markdown("---")

# -----------------------------
# APP 1: Shopee & CPAS (original code wrapped into function)
# -----------------------------


def app_shopee_cpas():
    # --- Page config and CSS for Shopee theme (scoped to this page) ---
    st.title("Shopee & CPAS — Excel Utilities")

    st.markdown("""
    <style>
    /* Scoped Shopee style (applied only when this function runs) */
    html, body, [data-testid="stAppViewContainer"], .stApp { background-color: #ffffff !important; }
    h1,h2,h3,h4,h5,h6,p,label { color: #EE4C29 !important; }
    section[data-testid="stSidebar"] > div:first-child { background-color: #EE4C29 !important; }
    section[data-testid="stSidebar"] * { color: #ffffff !important; }
    header, div[role="banner"], [data-testid="stToolbar"] { background-color: #EE4C29 !important; color: #ffffff !important; }
    div[data-testid="stFileUploader"], div[data-testid="stDropzone"], .stFileUploader { background-color: #EE4C29 !important; color: #ffffff !important; border: 1px solid #EE4C29 !important; box-shadow: none !important; }
    div[data-testid="stFileUploader"] button, .stFileUploader .stButton>button { background-color: #ffffff !important; color: #EE4C29 !important; border: 1px solid #ffffff !important; }
    table.dataframe thead th, .stDataFrame thead th, .ag-theme-alpine .ag-header { background-color: #EE4C29 !important; color: #ffffff !important; }
    a, .stMarkdown a { color: #EE4C29 !important; }
    section[data-testid="stSidebar"] svg { fill: #ffffff !important; stroke: #ffffff !important; }
    
    /* Tambahan sedikit untuk menata gaya Tabs agar warnanya sesuai dengan CSS kamu */
    div[data-testid="stTabs"] button { color: #EE4C29 !important; font-weight: bold; }
    div[data-testid="stTabs"] button[aria-selected="true"] { border-bottom-color: #EE4C29 !important; }
    </style>
    """, unsafe_allow_html=True)

    # ==========================================
    # HELPER FUNCTIONS
    # ==========================================
    def read_uploaded_bytes(uploaded_file) -> Optional[bytes]:
        if uploaded_file is None:
            return None
        try:
            uploaded_file.seek(0)
        except Exception:
            pass
        return uploaded_file.read()

    def to_excel_bytes_from_sheets(sheets: dict) -> bytes:
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            for sheet_name, df in sheets.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                if "Ringkasan" in sheet_name:
                    try:
                        ws = writer.sheets[sheet_name]
                        for col_idx in range(1, len(df.columns) + 1):
                            col_letter = get_column_letter(col_idx)
                            ws.column_dimensions[col_letter].width = 40
                            cell = ws.cell(row=2, column=col_idx)
                            cell.alignment = Alignment(wrap_text=True, vertical="top")
                    except Exception:
                        pass
                        
        output.seek(0)
        return output.getvalue()

    def swap_dot_comma_df(df: pd.DataFrame) -> pd.DataFrame:
        def swap_cell(x):
            if isinstance(x, str):
                return x.replace('.', 'DOT').replace(',', '.').replace('DOT', ',')
            return x
        if hasattr(df, 'map'):
            return df.map(swap_cell)
        return df.applymap(swap_cell)

    @st.cache_data
    def load_uploaded_csv_bytes(file_bytes: bytes) -> pd.DataFrame:
        if file_bytes is None:
            raise ValueError("No file bytes provided")
        raw = file_bytes.decode("utf-8", errors="ignore")
        lines = raw.splitlines()

        HEADER_KEYS = ["Nama Iklan", "Nama Iklan/Produk"]
        header_idx = None
        for i, line in enumerate(lines[:30]):
            if any(k in line for k in HEADER_KEYS):
                header_idx = i
                break
        if header_idx is None:
            raise ValueError("Header Nama Iklan tidak ditemukan")

        delimiter = ";" if lines[header_idx].count(";") > lines[header_idx].count(",") else ","
        clean_csv = "\n".join(lines[header_idx:])
        df = pd.read_csv(io.StringIO(clean_csv), sep=delimiter, engine="python", on_bad_lines="skip")
        df.columns = df.columns.str.strip()
        return df

    def normalize_nama_iklan_column(df: pd.DataFrame) -> pd.DataFrame:
        for col in ["Nama Iklan", "Nama Iklan/Produk"]:
            if col in df.columns:
                return df.rename(columns={col: "Nama Iklan"})
        raise ValueError("Kolom Nama Iklan tidak ditemukan")

    def short_nama_iklan(nama, max_words=2):
        if pd.isna(nama): return nama
        text = str(nama).strip()
        if text.lower().startswith("grup"): return text.split(" - ")[0]
        text = re.sub(r"\[.*?\]", "", text).strip()

        feature_blacklist = {"gamis", "busui","friendly","bahan","soft","ultimate","ultimates","motif","size","ukuran","promo","diskon","broad","testing","rayon","katun","cotton","silk","sustra","viscose","linen","polyester","jersey","crepe","chiffon","woolpeach","baloteli","babyterry","pink","hitam","black","putih","white","navy","biru","blue","merah","red","hijau","green","coklat","brown","abu","abu-abu","grey","gray","cream","krem","beige","maroon","ungu","purple","tosca","olive","sage", "sale", "couple"}
        store_blacklist = {"official","shop","store","boutique","fashion","my","zahir","myzahir","by","original","premium"}
        context_blacklist = {"terbaru","new","update","launch","launching","viral","hits","best","seller","bestseller","kondangan","ramadhan","ramadan","harian","pesta","formal","casual","trend","trending","populer","2024","2025","2026","2027", "2028", "2029", "2030"}
        all_blacklists = feature_blacklist | store_blacklist | context_blacklist
        product_keywords = {"dress", "set", "reject", "lebaran", "tunik", "abaya", "blouse", "khimar", "rok", "pashmina", "hijab", "outer"}

        parts = re.split(r"\s*[-|,/]\s*", text)
        candidates = []
        for part in parts:
            words = part.split()
            valid_words = []
            for w in words:
                wl_clean = re.sub(r'[^a-z0-9]', '', w.lower())
                if wl_clean in all_blacklists or not wl_clean: continue
                valid_words.append(w)
            if valid_words: candidates.append(valid_words)

        best_candidate = []
        for cand in candidates:
            if len(cand) >= 2 and any(re.sub(r'[^a-z0-9]', '', w.lower()) in product_keywords for w in cand):
                best_candidate = cand

        if not best_candidate:
            for cand in candidates:
                if any(re.sub(r'[^a-z0-9]', '', w.lower()) in product_keywords for w in cand):
                    best_candidate = cand
                    break
        if not best_candidate:
            for cand in candidates:
                if len(cand) >= 2:
                    best_candidate = cand
                    break
        if not best_candidate and candidates: best_candidate = candidates[0]
        if not best_candidate: best_candidate = text.split()

        if len(best_candidate) > max_words:
            kw_idx = -1
            for i, w in enumerate(best_candidate):
                if re.sub(r'[^a-z0-9]', '', w.lower()) in product_keywords:
                    kw_idx = i
                    break
            if kw_idx != -1:
                start_idx = max(0, kw_idx - max_words + 1)
                if start_idx + max_words > len(best_candidate):
                    start_idx = max(0, len(best_candidate) - max_words)
                best_candidate = best_candidate[start_idx : start_idx + max_words]
            else:
                best_candidate = best_candidate[:max_words]

        return " ".join(best_candidate).title()

    def highlight_row(row):
        styles = [''] * len(row)
        roas = row.get('Efektifitas Iklan')
        sales = row.get('Produk Terjual')
        gmv = row.get('Penjualan Langsung (GMV Langsung)')
        cost = row.get('Biaya')

        if pd.isna(sales) or pd.isna(cost): return styles
        if (cost == 0) and (sales > 0): return ['color: #006400'] * len(row)
        if sales == 0 and cost >= 10000: return ['color: #FF0000'] * len(row)
        if sales == 0 and cost < 10000: return styles

        if pd.notna(roas):
            try:
                if roas < 8: styles = ['background-color: red'] * len(row)
                elif roas < 10: styles = ['background-color: yellow'] * len(row)
                else: styles = ['background-color: lightgreen'] * len(row)
            except Exception: pass

        try: nama_idx = row.index.get_loc('Nama Iklan')
        except Exception: nama_idx = None
        try: gmv_idx = row.index.get_loc('Penjualan Langsung (GMV Langsung)')
        except Exception: gmv_idx = None

        if sales > 0 and (pd.isna(gmv) or gmv == 0):
            if nama_idx is not None: styles[nama_idx] = 'background-color: lightblue'
            if gmv_idx is not None: styles[gmv_idx] = 'background-color: lightblue'
        return styles

    def get_iklan_color(row, csv_mode):
        roas = row.get('Efektifitas Iklan')
        sales = row.get('Produk Terjual')
        cost = row.get('Biaya')

        if pd.isna(sales) or pd.isna(cost): return None
        if (cost == 0) and (sales > 0): return None
        if sales == 0 and cost >= 10000: return None
        if sales == 0 and cost < 10000: return None

        if csv_mode == "CSV Grup Iklan (hanya iklan produk)":
            if pd.isna(roas): return "HIJAU" if sales > 0 else None

        if pd.isna(roas) or roas < 8: return "MERAH"
        elif roas < 10: return "KUNING"
        else: return "HIJAU"

    def normalize_cols(df):
        return df.rename(columns=lambda c: re.sub(r"\s+", " ", str(c).strip()))

    def drop_kode_variasi_cols(df):
        cols_to_drop = [c for c in df.columns if c.strip().lower() == "kode variasi"]
        return df.drop(columns=cols_to_drop, errors="ignore")

    def extract_variation_base(name):
        if pd.isna(name): return ""
        s = str(name).strip()
        if s == "" or s == "-": return ""
        if "," in s:
            parts = s.rsplit(",", 1)
            base = parts[0].strip()
        else:
            base = s
        return base

    def clean_idr_number(x):
        if isinstance(x, str):
            x = x.strip()
            if not x or x == '-': return 0.0
            x = x.replace('%', '')
            if ',' in x: x = x.replace('.', '').replace(',', '.')
            else: x = x.replace('.', '')
            return x
        return x

    def safe_div(a, b):
        try:
            a, b = float(a), float(b)
            return 0.0 if b == 0 else a / b
        except Exception: return 0.0

    def format_percentage(val):
        return f"{val * 100:.2f}%".replace('.', ',')

    def to_excel_bytes_with_styling(df, product_merge_col="Kode Produk", highlight_condition=None):
        buf = io.BytesIO()
        df.to_excel(buf, index=False, sheet_name="Sheet1")
        buf.seek(0)
        wb = load_workbook(buf)
        ws = wb.active

        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        prod_col_idx = header.index(product_merge_col) + 1 if product_merge_col in header else None

        idr_col_indices = []
        for i, col_name in enumerate(header):
            if col_name and "IDR" in str(col_name).upper():
                idr_col_indices.append(i + 1) 

        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        total_dropdown_fill = PatternFill(start_color="BDE2F5", end_color="BDE2F5", fill_type="solid")
        var_dropdown_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid") 
        grand_total_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid") 
        bold_font = Font(bold=True)

        last_col_idx = ws.max_column
        last_col_letter = get_column_letter(last_col_idx)

        dv = DataValidation(type="list", formula1='"Total,~"', allow_blank=True)
        ws.add_data_validation(dv)
        
        if ws.max_row > 2:
            dv.add(f"{last_col_letter}2:{last_col_letter}{ws.max_row - 1}")

        if prod_col_idx:
            start = 2
            while start <= ws.max_row:
                current = ws.cell(row=start, column=prod_col_idx).value
                if current == "Total": break
                end = start
                while end + 1 <= ws.max_row and ws.cell(row=end + 1, column=prod_col_idx).value == current:
                    end += 1
                if current is not None and start < end:
                    rng = get_column_letter(prod_col_idx) + str(start) + ":" + get_column_letter(prod_col_idx) + str(end)
                    ws.merge_cells(rng)
                start = end + 1

        if highlight_condition is not None:
            for i, row in df.iterrows():
                excel_row = i + 2
                
                if row.get("Kode Produk", "") == "Total":
                    for col in range(1, last_col_idx + 1):
                        cell = ws.cell(row=excel_row, column=col)
                        cell.fill = grand_total_fill
                        cell.font = bold_font
                    continue 

                is_total = False
                try: is_total = highlight_condition(row)
                except Exception: pass

                if is_total:
                    for col in range(1, last_col_idx): ws.cell(row=excel_row, column=col).fill = yellow_fill
                    ws.cell(row=excel_row, column=last_col_idx).fill = total_dropdown_fill
                else:
                    ws.cell(row=excel_row, column=last_col_idx).fill = var_dropdown_fill

        rupiah_format = '_-"Rp"* #,##0_-;-"Rp"* #,##0_-;_-"Rp"* "-"_-;_-@_-'
        for col_idx in idr_col_indices:
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 20 
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(row=r, column=col_idx)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = rupiah_format

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return out


    # =========================================================================
    # NAVIGATION VIA TABS (MENGGANTIKAN SIDEBAR)
    # =========================================================================
    tab_out, tab_analitik, tab_ads = st.tabs([
        "🗂️ Shopee Out Platform", 
        "✨ Analitik Produk", 
        "📊 Shopee Ads"
    ])

    # =========================================================================
    # FITUR 1: GABUNGAN CONVERT -> SORT -> FILTER
    # =========================================================================
    with tab_out:
        st.header("Gabungan: Convert Dot/Comma ➔ Sort ➔ Filter")
        st.write("Upload 1 file Excel. Proses akan berjalan otomatis dan menghasilkan 2 file Excel:")
        st.markdown("""
        * **File 1 (Converter)**: Seluruh sheet dari file asli ditukar titik & koma-nya.
        * **File 2 (Sort & Filter)**: Mengambil sheet **Performa Produk**, melakukan Sort, lalu difilter untuk nama produk Terjual & ATC. Dibuatkan juga Ringkasan Filter per Platform.
        """)

        uploaded = st.file_uploader("📂 Upload file Excel (.xlsx/.xls)", type=["xlsx", "xls"], key="gabung_uploader_shopee")
        if uploaded:
            data = read_uploaded_bytes(uploaded)
            base_name = uploaded.name.rsplit(".", 1)[0]
            
            try:
                xls = pd.ExcelFile(BytesIO(data))

                # TAHAP 1
                sheets_convert = {}
                for sheet_name in xls.sheet_names:
                    df_c = pd.read_excel(xls, sheet_name=sheet_name, dtype=str)
                    df_c = swap_dot_comma_df(df_c)
                    sheets_convert[sheet_name] = df_c
                excel_bytes_convert = to_excel_bytes_from_sheets(sheets_convert)

                # TAHAP 2
                target_sheet_sort = "Performa Produk" if "Performa Produk" in xls.sheet_names else xls.sheet_names[0]
                df_raw_sort = pd.read_excel(xls, sheet_name=target_sheet_sort)
                req_sort = ["Channel", "Kode Produk"]
                missing_sort = [c for c in req_sort if c not in df_raw_sort.columns]
                
                df_sorted = pd.DataFrame()
                if not missing_sort:
                    df_sorted = df_raw_sort.sort_values(by=["Channel", "Kode Produk"], ascending=[True, True])
                else:
                    st.warning(f"⚠️ Kolom Sort tidak lengkap {missing_sort} di sheet '{target_sheet_sort}'. Menggunakan data tanpa sort.")
                    df_sorted = df_raw_sort.copy()

                # TAHAP 3
                df_terjual = pd.DataFrame()
                df_atc = pd.DataFrame()
                req_filter = ["Channel", "Produk", "Produk.1", "Produk Ditambahkan ke Keranjang"]
                missing_filter = [c for c in req_filter if c not in df_sorted.columns]
                
                if not missing_filter:
                    df_filter = df_sorted.copy()
                    df_filter["Produk.1"] = pd.to_numeric(df_filter["Produk.1"], errors="coerce").fillna(0)
                    df_filter["Produk Ditambahkan ke Keranjang"] = pd.to_numeric(df_filter["Produk Ditambahkan ke Keranjang"], errors="coerce").fillna(0)

                    df_terjual = df_filter[df_filter["Produk.1"] > 0][["Channel", "Produk"]].drop_duplicates().sort_values(by=["Channel", "Produk"]).reset_index(drop=True)
                    df_atc = df_filter[df_filter["Produk Ditambahkan ke Keranjang"] > 0][["Channel", "Produk"]].drop_duplicates().sort_values(by=["Channel", "Produk"]).reset_index(drop=True)

                    def generate_ringkasan(df_source):
                        res = {"Sales": [], "Traffic": [], "Instagram": []}
                        if not df_source.empty:
                            for _, r in df_source.iterrows():
                                ch = str(r["Channel"]).lower()
                                prod_short = short_nama_iklan(r["Produk"], max_words=2)
                                if "sales" in ch: res["Sales"].append(prod_short)
                                elif "traffic" in ch: res["Traffic"].append(prod_short)
                                elif "ig" in ch or "instagram" in ch: res["Instagram"].append(prod_short)
                                else: res["Sales"].append(prod_short)
                                    
                        final_dict = {}
                        for k in ["Sales", "Traffic", "Instagram"]:
                            unique_items = list(dict.fromkeys(res[k]))
                            if unique_items:
                                final_dict[k] = " ".join([f"{n}," for n in unique_items])
                            else:
                                final_dict[k] = ""
                        return pd.DataFrame([final_dict])

                    df_ringkasan_terjual = generate_ringkasan(df_terjual)
                    df_ringkasan_atc = generate_ringkasan(df_atc)
                else:
                    st.warning(f"⚠️ Kolom Filter tidak lengkap {missing_filter}. Tahap Filter dilewati.")

                # SUSUN EXCEL 2
                sheets_sort_filter = {"1_Data_Sorted": df_sorted}
                if not df_terjual.empty: sheets_sort_filter["2_Produk_Terjual"] = df_terjual
                if not df_atc.empty: sheets_sort_filter["3_Nama_Produk_ATC"] = df_atc
                if not df_ringkasan_terjual.empty: sheets_sort_filter["4_Ringkasan_Terjual"] = df_ringkasan_terjual
                if not df_ringkasan_atc.empty: sheets_sort_filter["5_Ringkasan_ATC"] = df_ringkasan_atc
                
                excel_bytes_sort_filter = to_excel_bytes_from_sheets(sheets_sort_filter)

                # UI DOWNLOAD
                st.success("✅ Seluruh proses selesai! Silakan unduh file hasilnya di bawah ini:")
                col1, col2 = st.columns(2)
                with col1:
                    st.download_button(
                        label="⬇️ Download Excel 1 (Dot/Comma)",
                        data=excel_bytes_convert,
                        file_name=f"{base_name}_converted.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                with col2:
                    st.download_button(
                        label="⬇️ Download Excel 2 (Sort & Filter)",
                        data=excel_bytes_sort_filter,
                        file_name=f"{base_name}_filtered.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                
                st.subheader("Preview File 2 - Sorted Data (10 Baris Pertama)")
                st.dataframe(df_sorted.head(10), use_container_width=True)

            except Exception as e:
                st.error(f"❌ Terjadi error: {e}")


    # =========================================================================
    # FITUR 2: Analitik Produk
    # =========================================================================
    with tab_analitik:
        st.header("Rapikan file XLSX/CSV — Produk & Variasi")
        st.markdown(
            "Upload file .xlsx atau .csv lalu tekan **Process**. Hasil bisa diunduh sebagai XLSX yang sudah di-merge, diberi warna, memiliki dropdown warna khusus, serta baris **Grand Total** di akhir."
        )

        uploaded = st.file_uploader("Upload file (.xlsx or .csv)", type=["xlsx", "xls", "csv"], key="rapiin_variasi_shopee")

        if uploaded is not None:
            base_name = uploaded.name.rsplit(".", 1)[0]
            
            try:
                if uploaded.name.lower().endswith((".xlsx", ".xls")): df_raw = pd.read_excel(uploaded, dtype=object)
                else: df_raw = pd.read_csv(uploaded, dtype=object)
            except Exception as e:
                st.error(f"Gagal membaca file: {e}")
                st.stop()

            df_raw = normalize_cols(df_raw)

            st.subheader("Preview (data asli, beberapa baris)")
            st.dataframe(df_raw.head(200))

            if st.button("Process", key="process_variasi_shopee"):
                df = df_raw.copy()
                df = drop_kode_variasi_cols(df)

                numeric_cols_guess = [
                    "Pengunjung Produk (Kunjungan)", "Halaman Produk Dilihat", "Pengunjung Melihat Tanpa Membeli",
                    "Klik Pencarian", "Suka", "Pengunjung Produk (Menambahkan Produk ke Keranjang)",
                    "Dimasukkan ke Keranjang (Produk)", "Total Pembeli (Pesanan Dibuat)", "Produk (Pesanan Dibuat)",
                    "Total Penjualan (Pesanan Dibuat) (IDR)", "Total Pembeli (Pesanan Siap Dikirim)",
                    "Produk (Pesanan Siap Dikirim)", "Penjualan (Pesanan Siap Dikirim) (IDR)"
                ]
                rate_cols_config = {
                    "Tingkat Pengunjung Melihat Tanpa Membeli": ("Pengunjung Melihat Tanpa Membeli", "Pengunjung Produk (Kunjungan)"),
                    "Tingkat Konversi Produk Dimasukkan ke Keranjang": ("Pengunjung Produk (Menambahkan Produk ke Keranjang)", "Pengunjung Produk (Kunjungan)"),
                    "Tingkat Konversi (Pesanan yang Dibuat)": ("Total Pembeli (Pesanan Dibuat)", "Pengunjung Produk (Kunjungan)"),
                    "Tingkat Konversi (Pesanan Siap Dikirim)": ("Total Pembeli (Pesanan Siap Dikirim)", "Pengunjung Produk (Kunjungan)"),
                    "Tingkat Konversi (Pesanan Siap Dikirim dibagi Pesanan Dibuat)": ("Total Pembeli (Pesanan Siap Dikirim)", "Total Pembeli (Pesanan Dibuat)")
                }

                if "Kode Produk" not in df.columns or "Nama Variasi" not in df.columns:
                    st.error("File harus berisi kolom 'Kode Produk' dan 'Nama Variasi'.")
                    st.stop()

                df["__NamaVariasiRaw"] = df["Nama Variasi"].astype(object)
                df["NamaVariasiBase"] = df["Nama Variasi"].apply(extract_variation_base)
                df["__is_total_row"] = df["NamaVariasiBase"].fillna("").apply(lambda s: True if s == "" else False)

                product_order = []
                seen = set()
                for i, r in df.iterrows():
                    kp = r.get("Kode Produk")
                    if kp not in seen:
                        seen.add(kp)
                        product_order.append(kp)

                variation_mask = ~df["__is_total_row"]
                agg_numeric = {}
                for c in df.columns:
                    if c in numeric_cols_guess:
                        df[c] = df[c].apply(clean_idr_number)
                        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
                        agg_numeric[c] = "sum"

                other_keep = ["SKU Induk", "Produk"] + list(rate_cols_config.keys())
                agg_other = {c: "first" for c in other_keep if c in df.columns}

                group_cols = ["Kode Produk", "NamaVariasiBase"]
                if variation_mask.any():
                    grouped = df[variation_mask].groupby(group_cols, dropna=False, as_index=False).agg({**agg_numeric, **agg_other})
                    grouped = grouped.rename(columns={"NamaVariasiBase": "Nama Variasi"})
                else:
                    grouped = pd.DataFrame(columns=["Kode Produk", "Nama Variasi"] + list(agg_numeric.keys()) + list(agg_other.keys()))

                totals = []
                for kp in product_order:
                    totals_rows = df[(df["Kode Produk"] == kp) & (df["__is_total_row"])]
                    if not totals_rows.empty:
                        tot = {"Kode Produk": kp}
                        for c in df.columns:
                            if c in other_keep: tot[c] = totals_rows.iloc[0].get(c)
                        for c in agg_numeric.keys():
                            tot[c] = totals_rows[c].astype(float).sum()
                        tot["Nama Variasi"] = ""
                        totals.append(pd.Series(tot))
                    else:
                        gi = grouped[grouped["Kode Produk"] == kp]
                        if not gi.empty:
                            tot = {"Kode Produk": kp, "Nama Variasi": ""}
                            for c in agg_numeric.keys(): tot[c] = gi[c].sum()
                            for c in other_keep:
                                any_row = df[df["Kode Produk"] == kp]
                                if not any_row.empty: tot[c] = any_row.iloc[0].get(c)
                            totals.append(pd.Series(tot))
                        else:
                            any_row = df[df["Kode Produk"] == kp]
                            if not any_row.empty:
                                row0 = any_row.iloc[0].copy()
                                row0["Nama Variasi"] = ""
                                totals.append(row0)

                totals_df = pd.DataFrame(totals).reset_index(drop=True)
                sort_col_induk = "Penjualan (Pesanan Siap Dikirim) (IDR)"
                if sort_col_induk in totals_df.columns:
                    totals_df[sort_col_induk] = pd.to_numeric(totals_df[sort_col_induk], errors="coerce").fillna(0)
                    totals_df = totals_df.sort_values(by=sort_col_induk, ascending=False)
                    
                product_order = totals_df["Kode Produk"].tolist()
                final_rows = []
                for kp in product_order:
                    tot_row = totals_df[totals_df["Kode Produk"] == kp]
                    if not tot_row.empty:
                        tot_row = tot_row.iloc[0].to_dict()
                        final_rows.append(tot_row)
                    
                    var_rows = grouped[grouped["Kode Produk"] == kp].copy()
                    if sort_col_induk in var_rows.columns:
                        var_rows[sort_col_induk] = pd.to_numeric(var_rows[sort_col_induk], errors="coerce").fillna(0)
                        var_rows = var_rows.sort_values(by=sort_col_induk, ascending=False)
                    
                    for _, vr in var_rows.iterrows():
                        final_rows.append(vr.to_dict())

                df_final = pd.DataFrame(final_rows).fillna("")

                for rate_col, (num_col, den_col) in rate_cols_config.items():
                    if num_col in df_final.columns and den_col in df_final.columns:
                        df_final[rate_col] = df_final.apply(lambda r: format_percentage(safe_div(r.get(num_col, 0), r.get(den_col, 0))), axis=1)

                def highlight_cond(row):
                    nv = row.get("Nama Variasi", "")
                    return (nv == "-" or str(nv).strip() == "")

                df_final["Nama Variasi"] = df_final["Nama Variasi"].replace({"": "-"})

                final_cols = []
                for c in df.columns:
                    if c == "Nama Variasi": continue 
                    if c in df_final.columns:
                        final_cols.append(c)
                        if c == "Produk": final_cols.append("Nama Variasi")
                            
                if "Nama Variasi" not in final_cols:
                    if "Kode Produk" in final_cols:
                        idx = final_cols.index("Kode Produk") + 1
                        final_cols.insert(idx, "Nama Variasi")
                    else:
                        final_cols.insert(0, "Nama Variasi")
                        
                for c in df_final.columns:
                    if c not in final_cols and not c.startswith("__"): final_cols.append(c)

                if "Tipe Baris" in final_cols: final_cols.remove("Tipe Baris")

                df_final["Tipe Baris"] = df_final.apply(lambda r: "Total" if highlight_cond(r) else "~", axis=1)
                final_cols.append("Tipe Baris")
                df_final = df_final[final_cols]

                total_rows_only = df_final[df_final["Tipe Baris"] == "Total"]
                grand_total_data = {}
                for c in final_cols:
                    if c == "Kode Produk": grand_total_data[c] = "Total"
                    elif c in numeric_cols_guess: grand_total_data[c] = pd.to_numeric(total_rows_only[c], errors="coerce").fillna(0).sum()
                    else: grand_total_data[c] = "-"
                
                df_final = pd.concat([df_final, pd.DataFrame([grand_total_data])], ignore_index=True)

                st.subheader("Hasil yang diproses (preview)")
                st.dataframe(df_final.tail(50)) 

                excel_bytes = to_excel_bytes_with_styling(df_final, product_merge_col="Kode Produk", highlight_condition=highlight_cond)

                st.download_button(
                    label="Unduh hasil (.xlsx, sudah merge, highlight, & Grand Total)",
                    data=excel_bytes,
                    file_name=f"{base_name}_sorted.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_rapi_xlsx_shopee"
                )

                csv_buf = io.BytesIO()
                csv_buf.write(df_final.to_csv(index=False).encode("utf-8"))
                csv_buf.seek(0)
                st.download_button(
                    label="Unduh hasil (.csv)",
                    data=csv_buf,
                    file_name=f"{base_name}_sorted.csv",
                    mime="text/csv",
                    key="dl_rapi_csv_shopee"
                )
                st.success("Selesai. Silakan unduh file atau cek pratinjau di atas.")


    # =========================================================================
    # FITUR 3: CSV IKLAN -> EXCEL BERWARNA
    # =========================================================================
    with tab_ads:
        st.header("Shopee Ads - CSV to Excel")
        st.markdown("Upload CSV iklan Shopee → otomatis rapi → download Excel laporan")

        # Pindahkan opsi sidebar ke dalam area ini
        st.markdown("##### Pengaturan Filter Laporan")
        csv_mode = st.selectbox(
            "Mode CSV",
            options=["CSV Keseluruhan (Normal)", "CSV Grup Iklan (hanya iklan produk)"],
            index=0,
            key="shopee_csv_mode_main"
        )
        
        st.markdown("Pilih kategori warna yang ingin disertakan di **RINGKASAN_IKLAN**")
        col1, col2, col3, col4 = st.columns(4)
        with col1: include_merah = st.checkbox("Sertakan MERAH", value=True, key="inc_merah_main")
        with col2: include_kuning = st.checkbox("Sertakan KUNING", value=True, key="inc_kuning_main")
        with col3: include_hijau = st.checkbox("Sertakan HIJAU", value=True, key="inc_hijau_main")
        with col4: include_biru = st.checkbox("Sertakan BIRU", value=True, key="inc_biru_main")
        
        st.caption("Catatan: filter warna ini hanya mempengaruhi sheet RINGKASAN_IKLAN (preview & export).")
        st.markdown("---")

        uploaded_file = st.file_uploader("Upload file CSV iklan Shopee", type=["csv"], key="csviklan_uploader_shopee")

        if uploaded_file:
            if st.button("🚀 Proses & Download Excel", key="process_csviklan_shopee"):
                try:
                    with st.spinner("Memproses data..."):
                        raw_bytes = read_uploaded_bytes(uploaded_file)
                        df = load_uploaded_csv_bytes(raw_bytes)
                        df = normalize_nama_iklan_column(df)

                        df["IS_AGGREGATE"] = df["Nama Iklan"].astype(str).str.lower().str.match(r'^\s*grup\b')

                        for col in ["Efektifitas Iklan", "Produk Terjual", "Penjualan Langsung (GMV Langsung)", "Biaya"]:
                            if col in df.columns:
                                df[col] = pd.to_numeric(df[col], errors="coerce")

                        df["IS_HIJAU_TIPE_A"] = (df.get("Biaya").notna() & (df.get("Biaya") == 0) & (df.get("Produk Terjual") > 0))
                        df["IS_BIRU"] = ((df.get("Produk Terjual", 0) > 0) & (df.get("Penjualan Langsung (GMV Langsung)", 0) == 0))
                        df["Nama Ringkasan"] = df["Nama Iklan"].where(df["IS_AGGREGATE"], df["Nama Iklan"].apply(short_nama_iklan))
                        df["Kategori"] = df.apply(lambda row: get_iklan_color(row, csv_mode), axis=1)

                        if csv_mode == "CSV Grup Iklan (hanya iklan produk)":
                            df_agg = df[df["IS_AGGREGATE"]].copy()
                            df_non_agg = df[~df["IS_AGGREGATE"]].copy()
                            df = pd.concat([df_non_agg, df_agg], ignore_index=True)

                            urutan_col = None
                            for c in df.columns:
                                if str(c).strip().lower() in ["urutan", "no", "no."]:
                                    urutan_col = c
                                    break
                            
                            if urutan_col:
                                new_vals = list(range(1, len(df_non_agg) + 1)) + [""] * len(df_agg)
                                df[urutan_col] = new_vals

                        if csv_mode == "CSV Grup Iklan (hanya iklan produk)":
                            df_nonagg = df[~df["IS_AGGREGATE"]].copy()
                        else:
                            df_nonagg = df.copy()

                        df_nonagg = df_nonagg[~df_nonagg["IS_HIJAU_TIPE_A"]].copy()

                        ordered_for_numbering = []
                        for _, row in df_nonagg.iterrows():
                            kat = row.get("Kategori")
                            if pd.notna(kat):
                                ordered_for_numbering.append({"nama": row["Nama Ringkasan"], "kategori": kat})
                            if row.get("IS_BIRU", False):
                                ordered_for_numbering.append({"nama": row["Nama Ringkasan"], "kategori": "BIRU"})

                        per_col = {"MERAH": [], "KUNING": [], "HIJAU": [], "BIRU": []}
                        if csv_mode != "CSV Keseluruhan (Normal)":
                            for kat in ["MERAH", "KUNING", "HIJAU"]:
                                names = df_nonagg[df_nonagg["Kategori"] == kat]["Nama Ringkasan"].tolist()
                                names = list(dict.fromkeys(names)) 
                                per_col[kat] = [f"{n}," for n in names]
                            
                            names_biru = df_nonagg[df_nonagg["IS_BIRU"]]["Nama Ringkasan"].tolist()
                            names_biru = list(dict.fromkeys(names_biru))
                            per_col["BIRU"] = [f"{n}," for n in names_biru]

                        tanpa_konversi_df = (
                            df_nonagg[(df_nonagg.get("Produk Terjual", 0) == 0) & (df_nonagg.get("Biaya", 0) >= 10000)]
                            [["Nama Ringkasan", "Biaya"]]
                            .rename(columns={"Nama Ringkasan": "Nama Iklan"})
                            .sort_values("Biaya", ascending=False)
                        )

                        hijau_cols = ["Nama Ringkasan", "Produk Terjual", "Efektifitas Iklan", "Biaya"]
                        available_cols = [c for c in hijau_cols if c in df.columns]
                        hijau_tipe_a_df = df[(df.get("Biaya").notna()) & (df.get("Biaya") == 0) & (df.get("Produk Terjual", 0) > 0)][available_cols].copy()
                        if "Nama Ringkasan" in hijau_tipe_a_df.columns:
                            hijau_tipe_a_df = hijau_tipe_a_df.rename(columns={"Nama Ringkasan": "Nama Iklan"})

                        filtered_per_col = {"MERAH": [], "KUNING": [], "HIJAU": [], "BIRU": []}
                        if include_merah: filtered_per_col["MERAH"] = per_col["MERAH"]
                        if include_kuning: filtered_per_col["KUNING"] = per_col["KUNING"]
                        if include_hijau: filtered_per_col["HIJAU"] = per_col["HIJAU"]
                        if include_biru: filtered_per_col["BIRU"] = per_col["BIRU"]

                        # EXPORT
                        buffer = io.BytesIO()
                        original_name = uploaded_file.name
                        base_name = original_name.rsplit(".", 1)[0]
                        filename = f"{base_name}_colored.xlsx"

                        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                            try:
                                styled = df.style.apply(highlight_row, axis=1)
                                styled.to_excel(writer, sheet_name="DATA_IKLAN", index=False)
                            except Exception:
                                df.to_excel(writer, sheet_name="DATA_IKLAN", index=False)

                            wb = writer.book
                            if "RINGKASAN_IKLAN" in wb.sheetnames:
                                wb.remove(wb["RINGKASAN_IKLAN"])
                            ws_ring = wb.create_sheet("RINGKASAN_IKLAN")

                            if csv_mode == "CSV Keseluruhan (Normal)":
                                ws_ring.cell(row=1, column=1, value="DAFTAR IKLAN (URUT)")
                                ws_ring.cell(row=1, column=1).font = Font(bold=True)
                                
                                semua_nama = []
                                for item in ordered_for_numbering:
                                    kat = item["kategori"]
                                    if (kat == "MERAH" and include_merah) or \
                                       (kat == "KUNING" and include_kuning) or \
                                       (kat == "HIJAU" and include_hijau) or \
                                       (kat == "BIRU" and include_biru):
                                        semua_nama.append(item["nama"])
                                
                                semua_nama = list(dict.fromkeys(semua_nama))
                                
                                if semua_nama:
                                    text_gabungan = "\n".join([f"{i+1}. {nama}" for i, nama in enumerate(semua_nama)])
                                    cell = ws_ring.cell(row=2, column=1, value=text_gabungan)
                                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                                    cell.font = Font(color="000000")
                                
                                ws_ring.column_dimensions["A"].width = 60
                                
                            else:
                                headers = ["MERAH", "KUNING", "HIJAU", "BIRU"]
                                color_map = {"MERAH": "FF0000", "KUNING": "000000", "HIJAU": "00AA00", "BIRU": "0066CC"}

                                for c_idx, h in enumerate(headers, start=1):
                                    cell = ws_ring.cell(row=1, column=c_idx, value=h)
                                    cell.font = Font(bold=True)

                                for c_idx, key in enumerate(headers, start=1):
                                    items = filtered_per_col.get(key, [])
                                    if items:
                                        joined = " ".join(items)
                                        if not joined.strip().endswith(","): joined = joined + ","
                                        cell = ws_ring.cell(row=2, column=c_idx, value=joined)
                                        cell.font = Font(color=color_map[key])
                                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                                    else:
                                        ws_ring.cell(row=2, column=c_idx, value="")

                                for i in range(1, 5):
                                    col_letter = get_column_letter(i)
                                    ws_ring.column_dimensions[col_letter].width = 40

                            tanpa_konversi_df.to_excel(writer, sheet_name=">10K_TANPA_KONVERSI", index=False)
                            ws_tc = writer.book[">10K_TANPA_KONVERSI"]
                            for r in range(2, ws_tc.max_row + 1):
                                for c in range(1, ws_tc.max_column + 1):
                                    cell = ws_tc.cell(row=r, column=c)
                                    cell.font = Font(color="FF0000")

                            hijau_tipe_a_df.to_excel(writer, sheet_name="SALES_0_BIAYA", index=False)
                            ws_hi = writer.book["SALES_0_BIAYA"]
                            for r in range(2, ws_hi.max_row + 1):
                                for c in range(1, ws_hi.max_column + 1):
                                    cell = ws_hi.cell(row=r, column=c)
                                    cell.font = Font(color="006400")

                        buffer.seek(0)

                    st.success("Excel laporan siap di-download 👇")
                    st.download_button(
                        "⬇️ Download Excel Laporan",
                        buffer,
                        filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="download_shopee_report"
                    )
                except Exception as e:
                    st.error(f"Terjadi error saat memproses file: {e}")
# -----------------------------
# APP 2: META KPI Highlight (wrapped)
# -----------------------------

def app_meta():
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter

    st.title("META Ads KPI Highlighter")

    st.markdown(
        """
        <style>
        /* Scoped META styling */
        html, body, .stApp, .reportview-container, .main, .block-container { background-color: #0066E7 !important; }
        section[data-testid="stSidebar"] > div:first-child { background-color: #0066E7 !important; }
        section[data-testid="stSidebar"] * { color: #ffffff !important; }
        div[data-testid="stFileUploader"] .upload-container,
        .stFileUploader > div {
            background-color: #ffffff !important;
            color: #0066E7 !important;
            border-radius: 8px !important;
            border: 1px solid rgba(0,102,231,0.18) !important;
        }
        .stFileUploader p, .stFileUploader label, .stFileUploader span { color: #0066E7 !important; }
        .stFileUploader button { background-color: #ffffff !important; color: #0066E7 !important; border: 1px solid #0066E7 !important; }
        .stDataFrame, .stDataFrame table, .ag-root { background-color: #ffffff !important; color: #000000 !important; }
        div[data-testid="stTabs"] button { color: #ffffff !important; font-weight: bold; }
        div[data-testid="stTabs"] button[aria-selected="true"] { color: #FFD700 !important; border-bottom-color: #FFD700 !important; }
        </style>
        """,
        unsafe_allow_html=True,
    )

    def is_number(x):
        try:
            if pd.isna(x): return False
            float(x)
            return True
        except:
            return False

    KEEP_DECIMAL_COLS = ["Frekuensi", "Tingkat klik tayang outbound"]

    tab_lama, tab_baru = st.tabs(["CPAS", "Whatsapp Ads"])

    # TAB 1: APLIKASI LAMA (STANDAR)
    with tab_lama:
        uploaded_file_lama = st.file_uploader("Upload file Excel (.xlsx) - Standar", type=["xlsx"], key="meta_uploader_lama")

        def highlight_cells_lama(val, column):
            try: v = float(val)
            except: return ""

            if column == "CPM (Biaya Per 1.000 Tayangan)" and v > 15000: return "background-color: #ffc7ce"
            if column == "CTR (Rasio Klik Tayang Tautan)" and v < 0.5: return "background-color: #ffc7ce"
            if column == "Frekuensi" and v > 3: return "background-color: #ffc7ce"
            if column == "ROAS Pembelian Khusus untuk Item Bersama" and v >= 10: return "background-color: #c6efce"
            return ""

        def format_cells_for_preview_lama(val, column):
            if pd.isna(val): return ""
            try: v = float(val)
            except: return val
            
            if "%ATC" in str(column):
                if v <= 1: v = v * 100
                return f"{v:.2f}%"
            
            if column in KEEP_DECIMAL_COLS: 
                return f"{v:.2f}"
            return f"{v:.0f}"

        def excel_highlight_and_write_lama(df):
            wb = Workbook()
            ws = wb.active
            ws.title = "KPI Highlight"

            for c_idx, col in enumerate(df.columns, start=1):
                ws.cell(row=1, column=c_idx, value=col)

            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

            for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
                for c_idx, col in enumerate(df.columns, start=1):
                    raw_val = row[col]
                    cell = ws.cell(row=r_idx, column=c_idx)

                    if is_number(raw_val):
                        v = float(raw_val)
                        if "%ATC" in str(col):
                            cell.value = v / 100.0 if v > 1 else v
                            cell.number_format = "0.00%"
                        elif col in KEEP_DECIMAL_COLS:
                            cell.value = v
                            cell.number_format = "0.##" 
                        else:
                            cell.value = v
                            cell.number_format = "0" 

                        try: eval_v = float(raw_val)
                        except: eval_v = None

                        if col == "CPM (Biaya Per 1.000 Tayangan)" and eval_v is not None and eval_v > 15000: cell.fill = red_fill
                        if col == "CTR (Rasio Klik Tayang Tautan)" and eval_v is not None and eval_v < 0.5: cell.fill = red_fill
                        if col == "Frekuensi" and eval_v is not None and eval_v > 3: cell.fill = red_fill
                        if col == "ROAS Pembelian Khusus untuk Item Bersama" and eval_v is not None and eval_v >= 10: cell.fill = green_fill
                    else:
                        cell.value = raw_val

            for i, col in enumerate(df.columns, start=1):
                ws.column_dimensions[get_column_letter(i)].width = min(max(15, len(str(col)) + 2), 50)

            out = BytesIO()
            wb.save(out)
            out.seek(0)
            return out

        if uploaded_file_lama:
            try:
                df_lama = pd.read_excel(uploaded_file_lama, header=0) 
                
                # Mendapatkan nama original (tanpa ekstensi)
                base_name_lama = uploaded_file_lama.name.rsplit(".", 1)[0]
                
                # Mengambil isi dari kolom "Awal pelaporan" jika ada
                tgl_awal_lama = ""
                if "Awal pelaporan" in df_lama.columns and not df_lama["Awal pelaporan"].dropna().empty:
                    raw_tgl = df_lama["Awal pelaporan"].dropna().iloc[0]
                    if pd.notna(raw_tgl):
                        # Jika format datetime, ubah jadi string YYYY-MM-DD. Jika bukan, ambil teksnya & hilangkan karakter ilegal /
                        tgl_awal_lama = raw_tgl.strftime("%Y-%m-%d") if hasattr(raw_tgl, 'strftime') else str(raw_tgl).replace("/", "-")
                
                # Bentuk nama file final
                if tgl_awal_lama:
                    final_filename_lama = f"{base_name_lama}_{tgl_awal_lama}_sorted.xlsx"
                else:
                    final_filename_lama = f"{base_name_lama}_sorted.xlsx"

                num_cols = df_lama.select_dtypes(include="number").columns
                df_lama[num_cols] = df_lama[num_cols].fillna(0)

                styled_df_lama = df_lama.style.apply(lambda col: [highlight_cells_lama(v, col.name) for v in col], axis=0)
                for col in df_lama.columns:
                    styled_df_lama = styled_df_lama.format(lambda v, c=col: format_cells_for_preview_lama(v, c), subset=[col])

                st.subheader("📌 Preview Data - Standar")
                st.dataframe(styled_df_lama, use_container_width=True)

                st.download_button(
                    label="⬇️ Download Excel (Standar)",
                    data=excel_highlight_and_write_lama(df_lama),
                    file_name=final_filename_lama,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_meta_lama"
                )
            except Exception as e:
                st.error(f"Gagal membaca file: {e}")

    # TAB 2: APLIKASI BARU (CUSTOM)
    with tab_baru:
        uploaded_file_baru = st.file_uploader("Upload file Excel (.xlsx) - Custom (Header Baris 3)", type=["xlsx"], key="meta_uploader_baru")

        def style_df_baru(df):
            styles = pd.DataFrame('', index=df.index, columns=df.columns)
            camp_col = next((c for c in df.columns if "kampanye" in str(c).lower() or "campaign" in str(c).lower()), None)

            for idx, row in df.iterrows():
                for col in df.columns:
                    val = row[col]
                    if is_number(val):
                        v = float(val)
                        if col == "CPM (Biaya Per 1.000 Tayangan)" and v > 15000: styles.loc[idx, col] = "background-color: #ffc7ce"
                        if col == "CTR (Rasio Klik Tayang Tautan)" and v < 0.5: styles.loc[idx, col] = "background-color: #ffc7ce"
                        if col == "Frekuensi" and v > 3: styles.loc[idx, col] = "background-color: #ffc7ce"
                            
                        if col == "Biaya per hasil" and camp_col is not None:
                            camp_name = str(row[camp_col]).lower() 
                            if "visit" in camp_name:
                                if v > 500:
                                    styles.loc[idx, col] = "background-color: #ffc7ce"
                            else:
                                if v > 5000:
                                    styles.loc[idx, col] = "background-color: #ffc7ce"
            return styles

        def format_cells_for_preview_baru(val, column):
            if pd.isna(val): return ""
            try: v = float(val)
            except: return val
            
            if "%ATC" in str(column):
                if v <= 1: v = v * 100
                return f"{v:.2f}%"
            
            if column in KEEP_DECIMAL_COLS: 
                return f"{v:.2f}"
            return f"{v:.0f}"

        def excel_highlight_and_write_baru(df):
            wb = Workbook()
            ws = wb.active
            ws.title = "KPI Highlight Custom"

            for c_idx, col in enumerate(df.columns, start=1):
                ws.cell(row=3, column=c_idx, value=col)

            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            camp_col = next((c for c in df.columns if "kampanye" in str(c).lower() or "campaign" in str(c).lower()), None)

            for r_idx, (_, row) in enumerate(df.iterrows(), start=4):
                for c_idx, col in enumerate(df.columns, start=1):
                    raw_val = row[col]
                    cell = ws.cell(row=r_idx, column=c_idx)

                    if is_number(raw_val):
                        v = float(raw_val)
                        if "%ATC" in str(col):
                            cell.value = v / 100.0 if v > 1 else v
                            cell.number_format = "0.00%"
                        elif col in KEEP_DECIMAL_COLS:
                            cell.value = v
                            cell.number_format = "0.##" 
                        else:
                            cell.value = v
                            cell.number_format = "0" 

                        try: eval_v = float(raw_val)
                        except: eval_v = None

                        if eval_v is not None:
                            if col == "CPM (Biaya Per 1.000 Tayangan)" and eval_v > 15000: cell.fill = red_fill
                            if col == "CTR (Rasio Klik Tayang Tautan)" and eval_v < 0.5: cell.fill = red_fill
                            if col == "Frekuensi" and eval_v > 3: cell.fill = red_fill
                            
                            if col == "Biaya per hasil" and camp_col is not None:
                                camp_name = str(row[camp_col]).lower()
                                if "visit" in camp_name:
                                    if eval_v > 500:
                                        cell.fill = red_fill
                                else:
                                    if eval_v > 5000:
                                        cell.fill = red_fill
                    else:
                        cell.value = raw_val

            for i, col in enumerate(df.columns, start=1):
                ws.column_dimensions[get_column_letter(i)].width = min(max(15, len(str(col)) + 2), 50)

            out = BytesIO()
            wb.save(out)
            out.seek(0)
            return out

        if uploaded_file_baru:
            try:
                df_baru = pd.read_excel(uploaded_file_baru, header=2) 
                
                # --- FITUR TAMBAHAN: Hapus kolom yang isinya kosong semua ---
                df_baru.dropna(axis=1, how='all', inplace=True)
                
                # Mendapatkan nama original (tanpa ekstensi)
                base_name_baru = uploaded_file_baru.name.rsplit(".", 1)[0]
                
                # Mengambil isi dari kolom "Awal pelaporan" jika ada
                tgl_awal_baru = ""
                if "Awal pelaporan" in df_baru.columns and not df_baru["Awal pelaporan"].dropna().empty:
                    raw_tgl = df_baru["Awal pelaporan"].dropna().iloc[0]
                    if pd.notna(raw_tgl):
                        tgl_awal_baru = raw_tgl.strftime("%Y-%m-%d") if hasattr(raw_tgl, 'strftime') else str(raw_tgl).replace("/", "-")
                
                # Bentuk nama file final
                if tgl_awal_baru:
                    final_filename_baru = f"{base_name_baru}_{tgl_awal_baru}_sorted.xlsx"
                else:
                    final_filename_baru = f"{base_name_baru}_sorted.xlsx"

                num_cols = df_baru.select_dtypes(include="number").columns
                df_baru[num_cols] = df_baru[num_cols].fillna(0)

                styled_df_baru = df_baru.style.apply(style_df_baru, axis=None)
                for col in df_baru.columns:
                    styled_df_baru = styled_df_baru.format(lambda v, c=col: format_cells_for_preview_baru(v, c), subset=[col])

                st.subheader("📌 Preview Data - Custom")
                st.dataframe(styled_df_baru, use_container_width=True)

                st.download_button(
                    label="⬇️ Download Excel (Custom Biaya per hasil)",
                    data=excel_highlight_and_write_baru(df_baru),
                    file_name=final_filename_baru,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_meta_baru"
                )
            except Exception as e:
                st.error(f"Gagal membaca file: {e}. Pastikan header tabel berada tepat di baris ke-3 Excel Anda.")               


# -----------------------------
# APP 3: TikTok (wrapped)
# -----------------------------

import streamlit as st
import pandas as pd
import numpy as np
import io
from datetime import datetime, date
from collections import OrderedDict
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from pandas.io.formats.style import Styler

def app_tiktok():
    st.title("🎵 Excel Tools — TikTok")

    # Helper & Config 
    percent_cols = [
        'Tingkat klik iklan produk', 'Rasio konversi iklan', 'Rasio tayang video iklan 2 detik',
        'Rasio tayang video iklan 6 detik', 'Rasio tayang video iklan 25%', 'Rasio tayang video iklan 50%',
        'Rasio tayang video iklan 75%', 'Rasio tayang video iklan 100%'
    ]

    def find_column(df, keywords):
        kws = [k.lower() for k in keywords]
        for col in df.columns:
            low = str(col).lower()
            if any(kw in low for kw in kws):
                return col
        return None

    def series_to_numeric_like(df_col):
        s_orig = df_col.astype(str).fillna("").str.strip()
        had_pct = s_orig.str.contains("%")
        s = s_orig.copy()
        has_paren = s.str.startswith("(") & s.str.endswith(")")
        s = s.mask(has_paren, "-" + s.str[1:-1])
        s = s.str.replace("%", "", regex=False).str.replace(",", "", regex=False).str.replace(" ", "", regex=False).replace("", np.nan)
        numeric = pd.to_numeric(s, errors="coerce")
        numeric = numeric.where(~had_pct, numeric / 100.0)
        return numeric

    def make_highlighter(col_biaya, col_pendapatan, col_roi, col_status):
        def highlight_row(row):
            styles = [''] * len(row)
            idx = {c: i for i, c in enumerate(row.index)}

            def parse_val(val):
                try:
                    if pd.isna(val): return np.nan
                    if isinstance(val, (int, float, np.floating, np.integer)): return float(val)
                    s = str(val).strip()
                    if s == "": return np.nan
                    had_pct = "%" in s
                    if s.startswith("(") and s.endswith(")"): s = "-" + s[1:-1]
                    num = float(s.replace("%", "").replace(",", "").replace(" ", ""))
                    return num / 100.0 if had_pct else num
                except Exception:
                    return np.nan

            try:
                biaya_val = parse_val(row[col_biaya]) if col_biaya in row.index else np.nan
                pendapatan_val = parse_val(row[col_pendapatan]) if col_pendapatan in row.index else np.nan
                roi_val = parse_val(row[col_roi]) if col_roi in row.index else np.nan
            except Exception:
                return styles

            if col_status is not None and col_status in row.index:
                status_text = str(row[col_status]).strip().lower() if pd.notna(row[col_status]) else ""
                if status_text == "perlu otorisasi":
                    styles = ['background-color: #98f073'] * len(row)
                    if col_status in idx: styles[idx[col_status]] = 'background-color: #ff7979'
                    return styles

            if pd.isna(roi_val): return styles
            biaya_pos = (pd.notna(biaya_val) and biaya_val > 0)
            pendapatan_pos = (pd.notna(pendapatan_val) and pendapatan_val > 0)
            if not (biaya_pos or pendapatan_pos) or roi_val == 0: return styles
            if roi_val >= 10: return ['background-color: #00ff00'] * len(row)
            if roi_val < 10: return ['background-color: #ffff00'] * len(row)
            return styles
        return highlight_row

    try:
        import xlsxwriter
        EXCEL_ENGINE = "xlsxwriter"
    except ImportError:
        EXCEL_ENGINE = "openpyxl"

    @st.cache_data
    def load_excel_safe(file, sheet_name=0):
        try:
            file.seek(0)
            temp_df = pd.read_excel(file, sheet_name=sheet_name, nrows=0, engine="openpyxl")
            dtype_dict = {}
            target_col = None
            for col in temp_df.columns:
                if "id" in str(col).lower():
                    dtype_dict[col] = str
                    target_col = col
                    break
            file.seek(0)
            final_df = pd.read_excel(file, sheet_name=sheet_name, dtype=dtype_dict, engine="openpyxl")
            
            # Membersihkan koma menjadi titik (Fixer)
            for col in final_df.columns:
                if col == target_col: continue
                if final_df[col].dtype == "object":
                    try:
                        final_df[col] = final_df[col].astype(str).str.replace(',', '.', regex=False)
                        final_df[col] = pd.to_numeric(final_df[col], errors='ignore')
                    except Exception: pass
            return final_df, target_col
        except Exception:
            return None, None

    # NAVBAR MINI TIKTOK (Halaman disederhanakan)
    PAGES_TIKTOK = ["Fitur Utama", "Daily Ads Comparator"]
    if "page_tiktok" not in st.session_state:
        st.session_state.page_tiktok = PAGES_TIKTOK[0]
        
    # --- INISIALISASI KEY DINAMIS UNTUK RESET UPLOADER ---
    if "tiktok_uploader_key" not in st.session_state:
        st.session_state["tiktok_uploader_key"] = 0

    cols = st.columns(len(PAGES_TIKTOK), gap="small")
    for i, p in enumerate(PAGES_TIKTOK):
        with cols[i]:
            if st.button(p, key=f"tiktok_nav_{i}"):
                st.session_state.page_tiktok = p
    st.markdown("---")

    # =========================================================================
    # HALAMAN 1: GABUNGAN EXCEL FIXER & PEWARNAAN ROI
    # =========================================================================
    if st.session_state.page_tiktok == "Fitur Utama":
        st.header("🛠️ Excel Fixer & Pewarnaan ROI")
        st.markdown("Mengamankan **ID Campaign**, mengubah koma `,` menjadi titik `.`, dengan opsi pewarnaan ROI.")

        uploaded_file = st.file_uploader("Upload File Excel (.xlsx / .xls)", type=["xlsx", "xls"], key="uploader_merged_tiktok")

        if uploaded_file:
            base_name = uploaded_file.name.rsplit('.', 1)[0]
            
            # Switch Pewarnaan ROI
            use_roi_color = st.toggle("🎨 Aktifkan Pewarnaan ROI", value=False, help="Jika aktif, baris dengan ROI tinggi/rendah akan diberi warna.")

            if st.button("🚀 Proses & Download", key="process_merged_tiktok"):
                with st.spinner("Memproses file..."):
                    df_hasil, kolom_target = load_excel_safe(uploaded_file)

                    if df_hasil is None:
                        st.error("Gagal memproses file. Pastikan format file benar.")
                    else:
                        buffer = io.BytesIO()

                        # JIKA SWITCH PEWARNAAN AKTIF
                        if use_roi_color:
                            outname = f"{base_name}_colored.xlsx"
                            
                            col_biaya = find_column(df_hasil, ["biaya", "cost"])
                            col_pendapatan_kotor = find_column(df_hasil, ["pendapatan kotor", "pendapatan_kotor", "pendapatan", "gmv", "revenue"])
                            col_pendapatan_bruto = find_column(df_hasil, ["pendapatan bruto", "penghasilan bruto", "penghasilan_bruto", "bruto", "gross", "gross revenue"])
                            col_roi = find_column(df_hasil, ["roi"])
                            col_status = find_column(df_hasil, ["status"])

                            col_pendapatan_effective = None
                            pendapatan_computed_name = "__pendapatan_bruto_computed"
                            bruto_was_computed = False

                            if col_pendapatan_bruto:
                                col_pendapatan_effective = col_pendapatan_bruto
                            elif col_pendapatan_kotor:
                                bonus_keywords = ["bonus", "komisi", "tunjangan", "insentif", "incentive"]
                                if any(any(k in str(c).lower() for k in bonus_keywords) for c in df_hasil.columns):
                                    col_pendapatan_effective = pendapatan_computed_name
                                    bruto_was_computed = True
                                else:
                                    col_pendapatan_effective = col_pendapatan_kotor

                            missing = [m for m, cond in zip(["Biaya", "Pendapatan", "ROI"], [col_biaya, col_pendapatan_kotor or col_pendapatan_bruto, col_roi]) if not cond]
                            
                            if missing:
                                st.error(f"Kolom wajib tidak ditemukan: {', '.join(missing)}. Gagal mewarnai ROI.")
                                st.stop()

                            biaya_num = series_to_numeric_like(df_hasil[col_biaya])
                            pendapatan_for_deletion = series_to_numeric_like(df_hasil[col_pendapatan_kotor if col_pendapatan_kotor else col_pendapatan_bruto])
                            roi_num = series_to_numeric_like(df_hasil[col_roi])
                            
                            delete_mask = (biaya_num == 0) & (pendapatan_for_deletion == 0) & (roi_num == 0)
                            df_colored = df_hasil.loc[~delete_mask].copy()

                            pct_present = [c for c in percent_cols if c in df_colored.columns]
                            for c in pct_present: df_colored[c] = series_to_numeric_like(df_colored[c])

                            if bruto_was_computed:
                                base = series_to_numeric_like(df_colored[col_pendapatan_kotor]).fillna(0)
                                extras = pd.Series(0.0, index=df_colored.index)
                                for bcol in [c for c in df_colored.columns if any(k in str(c).lower() for k in ["bonus", "komisi", "tunjangan", "insentif", "incentive"])]:
                                    extras += series_to_numeric_like(df_colored[bcol]).fillna(0)
                                df_colored[pendapatan_computed_name] = base + extras
                                col_pendapatan_effective = pendapatan_computed_name

                            if col_pendapatan_effective is None: col_pendapatan_effective = col_pendapatan_kotor or col_pendapatan_bruto

                            highlighter = make_highlighter(col_biaya, col_pendapatan_effective, col_roi, col_status)
                            styled = df_colored.style.apply(highlighter, axis=1)

                            with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                                styled.to_excel(writer, sheet_name="DATA_COLORED", index=False)
                                df_hasil.to_excel(writer, sheet_name="DATA_ASLI", index=False)
                                ws = writer.sheets["DATA_COLORED"]
                                for col in pct_present:
                                    try:
                                        col_idx = df_colored.columns.get_loc(col) + 1
                                        for row_cells in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, max_row=ws.max_row):
                                            for cell in row_cells:
                                                if isinstance(cell.value, (int, float, complex)) and not isinstance(cell.value, bool):
                                                    cell.number_format = '0.00%'
                                    except Exception: pass
                                    
                            st.success("✅ File berhasil diproses (Fixer + Warna).")
                            st.dataframe(df_colored.head(10), use_container_width=True)

                        # JIKA SWITCH PEWARNAAN MATI (Normal Fixer)
                        else:
                            outname = f"{base_name}_sorted.xlsx"
                            
                            with pd.ExcelWriter(buffer, engine=EXCEL_ENGINE) as writer:
                                df_hasil.to_excel(writer, index=False, sheet_name="Sheet1")
                                try:
                                    worksheet = writer.sheets["Sheet1"]
                                    if EXCEL_ENGINE == "xlsxwriter":
                                        for i, col in enumerate(df_hasil.columns):
                                            worksheet.set_column(i, i, max(df_hasil[col].astype(str).map(len).max(), len(str(col))) + 2)
                                    else:
                                        from openpyxl.utils import get_column_letter
                                        for i, col in enumerate(df_hasil.columns, 1):
                                            worksheet.column_dimensions[get_column_letter(i)].width = max(df_hasil[col].astype(str).map(len).max(), len(str(col))) + 2
                                except Exception: pass
                                
                            st.success("✅ File berhasil diproses (Hanya Fixer).")
                            st.dataframe(df_hasil.head(10), use_container_width=True)

                        buffer.seek(0)
                        st.download_button("📥 Download Excel Hasil", buffer, outname, key="download_merged_tiktok")


    # =========================================================================
    # HALAMAN 2: DAILY ADS COMPARATOR
    # =========================================================================
    elif st.session_state.page_tiktok == "Daily Ads Comparator":
        st.header("Ads Performance Comparator — DAILY FOCUS")
        st.markdown("""
        Upload TikTok exports per hari (header row 3, data row 4). Cache akan otomatis menyimpan dan menggabungkan datanya.
        """)

        ALLOWED_METRICS = [
            "ID", "Produk", "Status", "GMV", "Produk terjual", "Pesanan", "GMV tab Toko",
            "Impresi daftar produk tab Toko", "Rasio klik-tayang shop tab", "GMV dari LIVE",
            "Impresi dari LIVE", "Rasio klik-tayang dari LIVE", "GMV dari video",
            "Impresi dari video", "Rasio klik-tayang dari video", "Impresi dari kartu produk",
            "Tayangan halaman dari kartu produk", "Tayangan halaman unik dari kartu produk",
            "Pembeli unik dari kartu produk", "Rasio klik-tayang dari kartu produk",
            "Persentase konversi dari kartu produk",
        ]
        MAX_CACHE = 14
        PERCENT_NAME_KEYWORDS = ["rasio", "rasio klik", "persentase", "konversi", "ctr", "ratio"]

        def read_date_from_a1(uploaded_file) -> date:
            try:
                data = uploaded_file.read() if hasattr(uploaded_file, "read") else uploaded_file
                wb = load_workbook(filename=io.BytesIO(data) if isinstance(data, bytes) else data, data_only=True)
                raw = wb.active["A1"].value
                if isinstance(raw, datetime): return raw.date()
                if isinstance(raw, date): return raw
                if isinstance(raw, (int, float)):
                    try: return datetime.fromordinal(datetime(1900, 1, 1).toordinal() + int(raw) - 2).date()
                    except Exception: return raw
                if isinstance(raw, str):
                    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%Y/%m/%d"):
                        try: return datetime.strptime(raw.strip(), fmt).date()
                        except Exception: pass
                    return raw.strip()
                return raw
            except Exception:
                return None

        def read_data_table(uploaded_file) -> pd.DataFrame:
            try:
                if hasattr(uploaded_file, "read"):
                    try: uploaded_file.seek(0)
                    except Exception: pass
                return pd.read_excel(uploaded_file, header=2, engine="openpyxl")
            except Exception:
                return pd.DataFrame()

        def normalize_and_filter_df(df: pd.DataFrame) -> pd.DataFrame:
            df.columns = [str(c).strip() for c in df.columns]
            df = df.reindex(columns=[c for c in ALLOWED_METRICS if c in df.columns])
            for s in ["ID", "Produk", "Status"]:
                if s in df.columns: df[s] = df[s].astype(str)
            for col in df.columns:
                if col in ("ID", "Produk", "Status"): continue
                is_percent = any(k in col.lower() for k in PERCENT_NAME_KEYWORDS)
                if df[col].dtype == object or is_percent:
                    def try_parse(x):
                        if pd.isna(x): return None
                        if isinstance(x, str):
                            v = x.strip().replace(',', '')
                            if v.endswith('%'):
                                try: return float(v.rstrip('%')) / 100.0
                                except Exception: return None
                            try: return float(v)
                            except Exception: return None
                        if isinstance(x, (int, float)):
                            if is_percent and x > 1: return float(x) / 100.0
                            return float(x)
                        return None
                    df[col] = df[col].apply(try_parse)
                df[col] = pd.to_numeric(df[col], errors='coerce') if col not in ("ID", "Produk", "Status") else df[col]
            return df

        def add_to_session_cache(date_val, df):
            date_key = str(date_val)
            if "tiktok_daily_datasets" not in st.session_state:
                st.session_state["tiktok_daily_datasets"] = OrderedDict()
            datasets = st.session_state["tiktok_daily_datasets"]
            datasets[date_key] = df
            while len(datasets) > MAX_CACHE: datasets.popitem(last=False)
            st.session_state["tiktok_daily_datasets"] = datasets

        def clear_cache(): st.session_state["tiktok_daily_datasets"] = OrderedDict()

        def remove_date_from_cache(date_key):
            if "tiktok_daily_datasets" in st.session_state and date_key in st.session_state["tiktok_daily_datasets"]:
                st.session_state["tiktok_daily_datasets"].pop(date_key)

        def build_daily_aggregate(datasets: OrderedDict) -> pd.DataFrame:
            if not datasets: return pd.DataFrame()
            frames = []
            for date_key, df in datasets.items():
                parsed = pd.to_datetime(date_key, errors='coerce')
                if pd.isna(parsed):
                    try: parsed = pd.to_datetime(str(date_key).split()[0], errors='coerce')
                    except Exception: parsed = None
                if pd.isna(parsed): continue
                
                df2 = df[[c for c in ALLOWED_METRICS if c in df.columns]].copy()
                numeric = df2.select_dtypes(include=['number']).columns.tolist()
                summed = df2[numeric].sum(axis=0) if numeric else pd.Series(dtype=float)
                summed = summed.to_frame().T
                summed['date'] = pd.to_datetime(parsed)
                frames.append(summed)

            if not frames: return pd.DataFrame()
            agg = pd.concat(frames, ignore_index=True).set_index('date')
            agg.index = pd.to_datetime(agg.index).date
            return agg.sort_index()

        def style_daily_aggregate(df: pd.DataFrame) -> Styler:
            if df.empty: return df
            numeric_cols = df.select_dtypes(include=['number']).columns.tolist()
            diffs = df[numeric_cols].diff()
            styles = pd.DataFrame('', index=df.index, columns=df.columns)
            for col in numeric_cols:
                for idx in df.index:
                    d = diffs.loc[idx, col]
                    if pd.notna(d):
                        if d > 0: styles.at[idx, col] = 'background-color: #b6f2c2'
                        elif d < 0: styles.at[idx, col] = 'background-color: #f5b7b1'
                        else: styles.at[idx, col] = 'background-color: white'

            def fmt(x, col=None):
                if pd.isna(x): return ""
                if col and any(k in col.lower() for k in PERCENT_NAME_KEYWORDS):
                    try: return f"{x:.2%}"
                    except Exception: return x
                else:
                    try: return f"{int(x):,}" if float(x).is_integer() else f"{x:,.2f}"
                    except Exception: return x

            return df.style.format({c: (lambda v, col=c: fmt(v, col)) for c in df.columns}).apply(lambda _: styles, axis=None)

        def build_product_sheets(datasets: OrderedDict) -> bytes:
            if not datasets: return None
            frames = []
            for date_key, df in datasets.items():
                parsed = pd.to_datetime(str(date_key).split('~')[0].strip(), errors='coerce')
                if pd.notna(parsed):
                    d = df.copy()
                    d['date'] = pd.to_datetime(parsed)
                    frames.append(d)
            if not frames: return None

            concat = pd.concat(frames, ignore_index=True, sort=False)
            if 'Produk' not in concat.columns: return None

            numeric_metrics = [c for c in ALLOWED_METRICS if c in concat.columns and c not in ('ID', 'Produk', 'Status')]
            bytes_io = io.BytesIO()
            with pd.ExcelWriter(bytes_io, engine='openpyxl') as writer:
                for product_name, grp in concat.groupby('Produk'):
                    row = grp.groupby('date')[numeric_metrics].sum().reset_index().sort_values('date')
                    safe_sheet_name = str(product_name)[:31] if product_name else 'Unknown'
                    row.to_excel(writer, sheet_name=safe_sheet_name, index=False)
                    ws = writer.book[safe_sheet_name]

                    ws.column_dimensions['A'].width = 15
                    for cell in ws['A'][1:]: cell.number_format = 'yyyy-mm-dd'

                    from openpyxl.formatting.rule import FormulaRule
                    from openpyxl.chart import LineChart, Reference

                    green_fill = PatternFill(start_color="B6F2C2", end_color="B6F2C2", fill_type="solid")
                    red_fill = PatternFill(start_color="F5B7B1", end_color="F5B7B1", fill_type="solid")

                    for col_idx in range(2, ws.max_column + 1):
                        col_name = str(ws.cell(row=1, column=col_idx).value).lower()
                        col_letter = ws.cell(row=1, column=col_idx).column_letter
                        is_percent = any(k in col_name for k in PERCENT_NAME_KEYWORDS)
                        for row_idx in range(2, ws.max_row + 1):
                            ws.cell(row=row_idx, column=col_idx).number_format = '0.00%' if is_percent else '#,##0'
                        if ws.max_row >= 3:
                            cf_range = f"{col_letter}3:{col_letter}{ws.max_row}"
                            ws.conditional_formatting.add(cf_range, FormulaRule(formula=[f"{col_letter}3>{col_letter}2"], fill=green_fill))
                            ws.conditional_formatting.add(cf_range, FormulaRule(formula=[f"{col_letter}3<{col_letter}2"], fill=red_fill))

                    if ws.max_row >= 2:
                        start_chart_row, chart_idx = ws.max_row + 3, 0
                        for col_idx in range(2, ws.max_column + 1):
                            chart = LineChart()
                            chart.title = ws.cell(row=1, column=col_idx).value
                            chart.style, chart.width, chart.height, chart.legend = 13, 16, 8, None
                            chart.add_data(Reference(ws, min_col=col_idx, min_row=1, max_row=ws.max_row), titles_from_data=True)
                            chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=ws.max_row))
                            ws.add_chart(chart, f"{'A' if chart_idx % 2 == 0 else 'I'}{start_chart_row + (chart_idx // 2) * 16}")
                            chart_idx += 1

            bytes_io.seek(0)
            return bytes_io.read()

        col1, col2 = st.columns([2, 1])

        with col1:
            # --- MENGGUNAKAN KEY DINAMIS AGAR BISA DIRESET ---
            uploaded_files = st.file_uploader(
                "Upload TikTok exports (Excel .xlsx)", 
                type=["xlsx"], 
                accept_multiple_files=True, 
                key=f"tiktok_daily_uploader_{st.session_state['tiktok_uploader_key']}"
            )
            
            sukses_tanggal = [] 
            if uploaded_files:
                for uploaded in uploaded_files:
                    uploaded_bytes = uploaded.read()
                    date_val = read_date_from_a1(io.BytesIO(uploaded_bytes))
                    if not date_val:
                        st.error(f"Gagal ekstrak tanggal dari file: {uploaded.name}")
                    else:
                        df_raw = read_data_table(io.BytesIO(uploaded_bytes))
                        if df_raw.empty:
                            st.error(f"Gagal baca data tabel: {uploaded.name}")
                        else:
                            add_to_session_cache(date_val, normalize_and_filter_df(df_raw))
                            sukses_tanggal.append(str(date_val))
            if sukses_tanggal:
                st.success(f"Berhasil menyimpan {len(sukses_tanggal)} dataset untuk tanggal: {', '.join(sukses_tanggal)}")

        with col2:
            datasets = st.session_state.get("tiktok_daily_datasets", OrderedDict())
            if not datasets:
                st.info("Cache kosong.")
            else:
                st.write("**Datasets in cache**")
                st.table(pd.DataFrame([{"date": k, "rows": len(v)} for k, v in datasets.items()]).set_index('date'))
                to_remove = st.selectbox("Hapus tanggal (pilih)", [""] + list(datasets.keys()), key="tiktok_daily_remove")
                
                # --- MENAMBAHKAN INCREMENT KEY SAAT HAPUS ---
                if to_remove and st.button("Hapus tanggal", key="tiktok_daily_btn_rem"):
                    remove_date_from_cache(to_remove)
                    st.session_state["tiktok_uploader_key"] += 1 # Reset Uploader UI
                    st.rerun()
                    
                # --- MENAMBAHKAN INCREMENT KEY SAAT CLEAR ALL ---
                if st.button("Clear all cache", key="tiktok_daily_btn_clr"):
                    clear_cache()
                    st.session_state["tiktok_uploader_key"] += 1 # Reset Uploader UI
                    st.rerun()

        st.markdown("---")
        if not datasets: st.stop()

        valid_dates = [pd.to_datetime(str(k).split('~')[0].strip(), errors='coerce').date() for k in datasets.keys()]
        valid_dates = sorted([d for d in valid_dates if pd.notna(d)])
        
        # Penamaan File Download
        if len(valid_dates) >= 1:
            start_date_str = valid_dates[0].strftime("%Y%m%d")
            end_date_str = valid_dates[-1].strftime("%Y%m%d")
            outname_compare = f"dailycompare_{start_date_str}_to_{end_date_str}.xlsx"
        else:
            outname_compare = "dailycompare_report.xlsx"

        if len(valid_dates) > 1:
            expected_days = (valid_dates[-1] - valid_dates[0]).days + 1
            if len(valid_dates) < expected_days:
                expected_set = {valid_dates[0] + pd.Timedelta(days=i) for i in range(expected_days)}
                missing_str = ", ".join([d.strftime("%Y-%m-%d") for d in sorted(expected_set - set(valid_dates))])
                st.warning(f"⚠️ **Peringatan Data Bolong!** Ada tanggal yang terlewat: {missing_str}")

        st.subheader("📥 Export Laporan Akhir")
        excel_bytes = build_product_sheets(datasets)
        
        if excel_bytes:
            st.download_button("Download Excel Laporan (1 Sheet per Produk + Grafik)", excel_bytes, outname_compare, mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', key="tiktok_daily_dl_excel")
        else:
            st.info("Unggah file yang memiliki kolom Produk untuk membuat format Excel per-sheet.")

        st.markdown("---")
        
        frames = []
        for date_key, df in datasets.items():
            parsed = pd.to_datetime(str(date_key).split('~')[0].strip(), errors='coerce')
            if pd.notna(parsed):
                d = df.copy()
                d['date'] = parsed.date()
                frames.append(d)
                
        if not frames: st.stop()
        all_data = pd.concat(frames, ignore_index=True, sort=False)
        numeric_metrics = [c for c in ALLOWED_METRICS if c in all_data.columns and c not in ('ID', 'Produk', 'Status')]
        daftar_produk = sorted([p for p in all_data['Produk'].unique() if str(p).strip() not in ('nan', '', 'None')]) if 'Produk' in all_data.columns else []

        def show_charts(df_plot):
            if df_plot.empty: return st.info("Data tidak cukup untuk grafik.")
            col_a, col_b = st.columns(2)
            for idx, metric in enumerate(numeric_metrics):
                if metric in df_plot.columns:
                    with (col_a if idx % 2 == 0 else col_b):
                        st.caption(f"**{metric}**")
                        st.line_chart(df_plot[[metric]])

        tabs = st.tabs(["📊 Keseluruhan (All)"] + [f"🛍️ {p[:20]}..." if len(p) > 20 else f"🛍️ {p}" for p in daftar_produk])
        
        with tabs[0]:
            agg = build_daily_aggregate(datasets)
            if agg.empty: st.warning("Tidak ada data numerik.")
            else:
                sub1, sub2 = st.tabs(["🧮 Tabel Data", "📈 Grafik Tren"])
                with sub1:
                    st.write(style_daily_aggregate(agg).to_html(), unsafe_allow_html=True)
                    st.download_button("📥 Download CSV (All)", agg.reset_index().to_csv(index=False), "daily_aggregate_all.csv", mime='text/csv', key="tiktok_daily_dl_csv")
                with sub2: show_charts(agg)

        for i, produk_name in enumerate(daftar_produk):
            with tabs[i + 1]:
                df_produk = all_data[all_data['Produk'] == produk_name]
                agg_produk = df_produk.groupby('date')[numeric_metrics].sum().sort_index()
                if agg_produk.empty: st.info("Tidak ada data numerik.")
                else:
                    sub1, sub2 = st.tabs(["🧮 Tabel Data", "📈 Grafik Tren"])
                    with sub1: st.write(style_daily_aggregate(agg_produk).to_html(), unsafe_allow_html=True)
                    with sub2: show_charts(agg_produk)


# -----------------------------
# MAIN: render navbar then the selected app
# -----------------------------

def main():
    st.sidebar.title("Multi-Platform Dashboard")
    st.sidebar.markdown("Pilih platform dari navbar atas atau dari sini:")
    
    # 4. Ikat selectbox langsung ke st.session_state.page menggunakan parameter 'key'
    # Tidak perlu lagi st.session_state.page = chosen, karena 'key="page"' 
    # akan otomatis mengubah st.session_state.page saat opsi dipilih.
    st.sidebar.selectbox(
        "Pilih platform (sidebar)", 
        options=PAGES, 
        key="page" 
    )

    # Render navbar atas
    navbar()

    # Routing ke aplikasi masing-masing
    if st.session_state.page == PAGES[0]:
        app_shopee_cpas()
    elif st.session_state.page == PAGES[1]:
        app_meta()
    else:
        app_tiktok()

if __name__ == "__main__":
    main()
